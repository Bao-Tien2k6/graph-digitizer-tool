"""
pipeline/coordinate_transform.py
=================================
Stage 5 — Coordinate Transform & Export
"""

from __future__ import annotations

import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional

import cv2
import matplotlib
matplotlib.use("Agg")
import matplotlib.patches as mpatches
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from pipeline.axes_detector import AxesInfo, ScaleType
from pipeline.parallel_router import DetectionResult
from pipeline.preprocess import BGRImage


# Public data structures
@dataclass
class DataPoint:
    x: float
    y: float
    delta_x: float = 0.0
    delta_y: float = 0.0
    series_id: int = 0
    pixel_x: float = 0.0
    pixel_y: float = 0.0


@dataclass
class TransformResult:
    chart_type: str
    points: List[DataPoint] = field(default_factory=list)
    affine_matrix: Optional[np.ndarray] = None
    output_paths: Dict[str, Path] = field(default_factory=dict)


def round_value(v: float) -> float | int:
    """Round an extracted value: whole numbers → int, otherwise 1 decimal place."""
    r = round(float(v), 3)
    return int(r) if r == int(r) else r


# Main entry point
def transform_and_export(
    detection: DetectionResult,
    axes: AxesInfo,
    original_img: BGRImage,
    output_dir: str | Path,
    formats: List[str] = ("csv", "xlsx", "overlay_png"),
    stem: str = "result",
) -> TransformResult:
    """Convert pixel detections to data values and write output files."""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    affine = _build_affine_matrix(axes)
    points = _pixel_to_data(detection.pixel_points, affine, axes)
    points = _propagate_uncertainty(points, axes, detection)

    result = TransformResult(
        chart_type=detection.chart_type,
        points=points,
        affine_matrix=affine,
    )

    for fmt in formats:
        if fmt == "csv":
            path = _export_csv(points, output_dir, stem)
        elif fmt == "json":
            path = _export_json(points, detection.chart_type, output_dir, stem)
        elif fmt == "xlsx":
            path = _export_xlsx(points, output_dir, stem)
        elif fmt == "overlay_png":
            path = _export_overlay_png(points, original_img, output_dir, stem)
        else:
            raise ValueError(f"Unsupported export format: {fmt!r}")
        result.output_paths[fmt] = path

    return result


# Step helpers
def _build_affine_matrix(axes: AxesInfo) -> np.ndarray:
    """Build 2×3 pixel→data affine matrix from axis scale coefficients."""
    ax, bx = axes.x_axis.scale_coeffs  # data_x = ax * pixel_x + bx
    ay, by = axes.y_axis.scale_coeffs  # data_y = ay * pixel_y + by
    return np.array([
        [ax, 0.0, bx],
        [0.0, ay, by],
    ], dtype=float)


def _pixel_to_data(
    pixel_points: List[Dict],
    affine: np.ndarray,
    axes: AxesInfo,
) -> List[DataPoint]:
    """Apply affine transform to pixel detections."""
    ax, bx = affine[0, 0], affine[0, 2]
    ay, by = affine[1, 1], affine[1, 2]

    points: List[DataPoint] = []
    for pp in pixel_points:
        px = float(pp.get("x", 0))
        py = float(pp.get("y", 0))
        series = int(pp.get("series", 0))

        if axes.x_axis.scale_type == ScaleType.LOG10:
            data_x = 10 ** (ax * px + bx)
        else:
            data_x = ax * px + bx

        if axes.y_axis.scale_type == ScaleType.LOG10:
            data_y = 10 ** (ay * py + by)
        else:
            data_y = ay * py + by

        points.append(DataPoint(
            x=round_value(data_x),
            y=round_value(data_y),
            series_id=series,
            pixel_x=px,
            pixel_y=py,
        ))

    return points


def _propagate_uncertainty(
    points: List[DataPoint],
    axes: AxesInfo,
    detection: DetectionResult,
) -> List[DataPoint]:
    """Estimate ±δ_x and ±δ_y for each data point."""
    ax = abs(axes.x_axis.scale_coeffs[0])
    ay = abs(axes.y_axis.scale_coeffs[0])

    for i, dp in enumerate(points):
        pp = detection.pixel_points[i] if i < len(detection.pixel_points) else {}
        area = float(pp.get("area", 50))

        # Detection localisation uncertainty (blob radius → data units)
        r_px = np.sqrt(area / np.pi)
        delta_detect_x = r_px * ax if ax > 0 else 0.0
        delta_detect_y = r_px * ay if ay > 0 else 0.0

        # Scale fit quality uncertainty
        if axes.x_axis.scale_r2 < 1.0 and ax > 0:
            ticks_x = [t.label_value for t in axes.x_axis.ticks if t.ocr_confidence > 0]
            x_range = (max(ticks_x) - min(ticks_x)) if len(ticks_x) > 1 else 1.0
            delta_fit_x = (1.0 - axes.x_axis.scale_r2) * x_range * 0.5
        else:
            delta_fit_x = 0.0

        if axes.y_axis.scale_r2 < 1.0 and ay > 0:
            ticks_y = [t.label_value for t in axes.y_axis.ticks if t.ocr_confidence > 0]
            y_range = (max(ticks_y) - min(ticks_y)) if len(ticks_y) > 1 else 1.0
            delta_fit_y = (1.0 - axes.y_axis.scale_r2) * y_range * 0.5
        else:
            delta_fit_y = 0.0

        dp.delta_x = round(np.sqrt(delta_detect_x ** 2 + delta_fit_x ** 2), 4)
        dp.delta_y = round(np.sqrt(delta_detect_y ** 2 + delta_fit_y ** 2), 4)

    return points


# Export functions
def _export_csv(points: List[DataPoint], output_dir: Path, stem: str) -> Path:
    """Write one CSV file per series."""
    from collections import defaultdict
    series_map: Dict[int, List[DataPoint]] = defaultdict(list)
    for p in points:
        series_map[p.series_id].append(p)

    first_path = output_dir / f"{stem}_series0.csv"
    if not series_map:
        # No points detected — still write a valid (header-only) file.
        first_path.write_text("x,y,delta_x,delta_y")
        return first_path

    for sid, pts in sorted(series_map.items()):
        path = output_dir / f"{stem}_series{sid}.csv"
        lines = ["x,y,delta_x,delta_y"]
        for p in sorted(pts, key=lambda p: p.x):
            lines.append(f"{p.x},{p.y},{p.delta_x},{p.delta_y}")
        path.write_text("\n".join(lines))
        if sid == 0:
            first_path = path

    return first_path


def _export_json(
    points: List[DataPoint],
    chart_type: str,
    output_dir: Path,
    stem: str,
) -> Path:
    """Write a single JSON file with all series."""
    from collections import defaultdict
    series_map: Dict[int, List[DataPoint]] = defaultdict(list)
    for p in points:
        series_map[p.series_id].append(p)

    data = {
        "chart_type": chart_type,
        "series": [
            {
                "id": sid,
                "points": [
                    {"x": p.x, "y": p.y, "delta_x": p.delta_x, "delta_y": p.delta_y}
                    for p in sorted(pts, key=lambda p: p.x)
                ],
            }
            for sid, pts in sorted(series_map.items())
        ],
    }
    path = output_dir / f"{stem}.json"
    path.write_text(json.dumps(data, indent=2))
    return path


def _export_xlsx(points: List[DataPoint], output_dir: Path, stem: str) -> Path:
    """Write an Excel workbook with one sheet per series (columns: x, y)."""
    from collections import defaultdict
    series_map: Dict[int, List[DataPoint]] = defaultdict(list)
    for p in points:
        series_map[p.series_id].append(p)

    wb = openpyxl.Workbook()

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEEFF")

    # A workbook must keep at least one visible sheet. When no points were
    # detected, leave a single header-only sheet rather than removing them all.
    if not series_map:
        ws = wb.active
        ws.title = "Series 0"
        for col, h in enumerate(["x", "y"], start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        path = output_dir / f"{stem}.xlsx"
        wb.save(str(path))
        return path

    wb.remove(wb.active)  # remove default sheet

    for sid, pts in sorted(series_map.items()):
        ws = wb.create_sheet(title=f"Series {sid}")
        headers = ["x", "y"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, p in enumerate(sorted(pts, key=lambda p: p.x), start=2):
            ws.cell(row=row_idx, column=1, value=p.x)
            ws.cell(row=row_idx, column=2, value=p.y)

    path = output_dir / f"{stem}.xlsx"
    wb.save(str(path))
    return path


def _export_overlay_png(
    points: List[DataPoint],
    original_img: BGRImage,
    output_dir: Path,
    stem: str,
) -> Path:
    """Render detected points on the original image and save as PNG."""
    # Convert BGR to RGB for matplotlib
    rgb = cv2.cvtColor(original_img, cv2.COLOR_BGR2RGB)

    palette = [
        "#e6194b", "#3cb44b", "#4363d8", "#f58231",
        "#911eb4", "#42d4f4", "#f032e6", "#bfef45",
    ]

    # Object-oriented Figure (no pyplot global state). Streamlit runs the app
    # in a worker thread where pyplot's shared figure manager is not
    # thread-safe; using Figure + FigureCanvasAgg directly avoids the crashes
    # that can accumulate across reruns.
    fig = Figure(figsize=(rgb.shape[1] / 100, rgb.shape[0] / 100), dpi=100)
    FigureCanvasAgg(fig)
    ax = fig.add_subplot(111)
    ax.imshow(rgb)
    ax.axis("off")

    from collections import defaultdict
    series_pts: Dict[int, List[DataPoint]] = defaultdict(list)
    for p in points:
        series_pts[p.series_id].append(p)

    legend_handles = []
    for sid, pts in sorted(series_pts.items()):
        color = palette[sid % len(palette)]
        xs = [p.pixel_x for p in pts]
        ys = [p.pixel_y for p in pts]
        ax.scatter(xs, ys, s=80, facecolors='none', edgecolors=color,
                   linewidths=2.0, zorder=5)
        # Crosshairs
        for p in pts:
            ax.plot([p.pixel_x - 6, p.pixel_x + 6], [p.pixel_y, p.pixel_y],
                    color=color, lw=0.8, zorder=4)
            ax.plot([p.pixel_x, p.pixel_x], [p.pixel_y - 6, p.pixel_y + 6],
                    color=color, lw=0.8, zorder=4)
            ax.annotate(f"({p.x:.2f}, {p.y:.2f})",
                        xy=(p.pixel_x, p.pixel_y),
                        xytext=(4, -12), textcoords="offset points",
                        fontsize=5, color=color, zorder=6)
        patch = mpatches.Patch(color=color, label=f"Series {sid}")
        legend_handles.append(patch)

    if legend_handles:
        ax.legend(handles=legend_handles, loc="upper right", fontsize=6,
                  framealpha=0.7)

    fig.tight_layout(pad=0)
    path = output_dir / f"{stem}_overlay.png"
    fig.savefig(str(path), dpi=150, bbox_inches="tight")
    return path
